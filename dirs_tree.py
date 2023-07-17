# -*- coding:UTF-8 -*- 

# author:
# contact: 
# datetime:
# software: PyCharm

import os
import os.path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN


s_top = 1  # >1
s_left = 1  # >1

color_map = {1: "CCFFCC", 2: "FFFF99", 3: "99CCFF", 4: "FF99CC", 5: "CC99FF", 6: "66FF99"}
CURR_COL = 0

workbook = Workbook()
sheet = workbook.active

# styles
bottom_border = Border(bottom=Side(border_style=BORDER_THIN, color='00000000'))
left_border = Border(left=Side(border_style=BORDER_THIN, color='00000000'))
top_border = Border(top=Side(border_style=BORDER_THIN, color='00000000'))
top_bottom_border = Border(
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
    )
top_left_border = Border(
    top=Side(border_style=BORDER_THIN, color='00000000'),
    left=Side(border_style=BORDER_THIN, color='00000000')
    )

# sheet.cell(row=3, column=2).border = bottom_border
text_alignment = Alignment(horizontal="center", vertical="top", text_rotation=180)
text_font = Font(bold=True, name="Verdana", size="10")



def dfs_showdir(path, depth):
    if depth == 0:
        print("root:[" + path + "]")

    for item in os.listdir(path):
        if item in ['.git', '.idea', '__pycache__']:
            continue

        new_item = os.path.join(path, item)
        if os.path.isdir(new_item):
            global CURR_COL
            CURR_COL += 1
            print("| " * depth + "+--" + item) # print folders tree
            global sheet
            sheet.cell(row=s_top+depth, column=s_left+CURR_COL).value = item
            sheet.merge_cells(start_row=s_top+depth, start_column=s_left+CURR_COL, end_row=s_top+9, end_column=s_left+CURR_COL)

            for i in range(1, depth):
                sheet.cell(row=s_top+i, column=s_left+CURR_COL).fill = PatternFill(start_color=color_map.get(i, "FFFFFF"), fill_type="solid")
                sheet.cell(row=s_top+i, column=s_left+CURR_COL).border = top_bottom_border

            sheet.cell(row=s_top+depth, column=s_left+CURR_COL).fill = PatternFill(start_color=color_map.get(depth, "FFFFFF"), fill_type="solid")
            sheet.cell(row=s_top+depth, column=s_left+CURR_COL).font = text_font
            sheet.cell(row=s_top+depth, column=s_left+CURR_COL).alignment = text_alignment
            for i in range(depth, 10):
                if depth == 1:
                    sheet.cell(row=s_top+i, column=s_left+CURR_COL).border = top_left_border
                else:
                    sheet.cell(row=s_top+i, column=s_left+CURR_COL).border = left_border
            dfs_showdir(new_item, depth + 1)


if __name__ == '__main__':
    dfs_showdir('.', 1)
    workbook.save(filename="folders_sheet.xlsx")
